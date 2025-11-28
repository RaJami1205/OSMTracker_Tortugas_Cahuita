import os
from openpyxl import Workbook, load_workbook

def append_to_excel(diccionario, ruta_excel="censo_tortugas.xlsx"):
    if not os.path.exists(ruta_excel):
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        ws.append(list(diccionario.keys()))
        wb.save(ruta_excel)
    wb = load_workbook(ruta_excel)
    ws = wb.active
    headers = list(diccionario.keys())
    excel_headers = [cell.value for cell in ws[1]]
    if headers != excel_headers:
        raise Exception(
            "Encabezados no coinciden con el Excel.\n"
            f"Excel: {excel_headers}\n"
            f"Datos: {headers}"
        )
    valores = []
    for h in headers:
        valor = diccionario[h]
        if isinstance(valor, list):
            valor = ", ".join(valor)
        valores.append(valor)
    ws.append(valores)
    wb.save(ruta_excel)
    print(" Datos agregados correctamente a", ruta_excel)

if __name__ == "__main__":
    # Crear diccionario de prueba si se ejecuta directamente
    diccionario_prueba = {
        "lat": 9.7489,
        "lon": -82.6751,
        "hora": "2025-11-28T10:30:00Z",
        "tipo_avistamiento": "Tortuga",
        "actividad": "Desovando",
        "especie": "Verde",
        "estado_pedunculo": "Completo",
        "direccion": "Mar",
        "zona_actividad": "Olas",
        "destino_nido": "In Situ",
        "causa_reubicacion": None,
        "tipo_vivero": None,
        "pit_tag": None,
        "hallazgos": ["Marcas", "Cicatrices"]
    }
    
    print(" Creando/Actualizando Excel de prueba...")
    append_to_excel(diccionario_prueba)
    print(" Proceso completado")
